import json
import locale
import os
import re
from datetime import datetime
from datetime import timezone, timedelta

import discord
import matplotlib.pyplot as plt
import numpy as np
import pytz
from discord import Option
from discord.ext import tasks, commands
from discord.ui import Select, View
from dotenv import load_dotenv
from loguru import logger
from wordcloud import WordCloud
import openpyxl
from openpyxl.styles import Font


from src.ft.bonus.squadbusters.navigation import NavigationView
from src.ft.ft1.recommendations import generate_recommendations
from src.ft.ft1.stream_notifications import check_streamers, validate_streamer
from src.ft.ft2.icals_to_json import register_user_ical
from src.ft.ft2.planning import is_everyone_available, download_ical, ensure_temp_dir, TEMP_DIR, parse_ical_content, \
    check_availability
from src.ft.ft2.weather import get_weather
from src.ft.ft3.profanities import handle_profanities
from src.ft.ft3.warnings import Warnings
from src.ft.ft4.gifs import handle_gifs_channel
from src.ft.ft5.gpt import GPT
from src.ft.ft5.reports import Reports
from src.utilities.settings import Settings
from src.utilities.utilities import setup_commands, get_current_date_formatted

load_dotenv()
DISCORD_BOT_TOKEN = os.getenv('DISCORD_BOT_TOKEN')
intents = discord.Intents.default()
intents.message_content = True
intents.members = True
bot = discord.Bot(intents=intents)
settings = Settings()
warnings = Warnings()
reports = Reports()

locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')


@bot.event
async def on_ready():
    """
    This function is an event handler that gets triggered when the bot is ready.
    This function doesn't take any arguments and doesn't return anything.
    """
    logger.success(f'Bot is ready. Logged in as {bot.user}')
    await handle_tasks()


async def handle_tasks():
    """
    Starts the scheduled tasks for the Discord bot.

    This function initializes and starts various scheduled tasks that the bot performs at set intervals. These tasks include:
    - Sending scheduled recommendations every hour.
    - Checking streamers' status and notifying the server accordingly.
    - Running a daily update task for maintaining current data.
    - Saving report data every minute to ensure data persistence.
    - Recommending activities based on the current weather every 24 hours with a 3-minute offset.

    Each task is started by calling the `.start()` method on the respective `tasks.loop` instance. The `bot` instance is
    passed as an argument to `check_streamers.start()` to enable it to access Discord server information.
    """
    # scheduled_recommendation.start()
    check_streamers.start(bot)
    scheduled_update.start()
    scheduled_reports_save.start()
    # scheduled_activity_recommendation.start()


@tasks.loop(hours=24)
async def scheduled_update():
    """
    This function is a task that runs every 24 hours.

    The function executes the "scraper.py" script located in the "src/ft/bonus/squadbusters" directory.
    The purpose of this script is to update the data used by the bot.

    This function doesn't take any arguments and doesn't return anything.
    """
    logger.info("Updating squadbusters data...")
    os.system("python src/ft/bonus/squadbusters/scraper.py")


@bot.event
async def on_message(message: discord.Message):
    """
    Handles incoming messages for various functionalities based on the message content and origin.

    This event handler performs several checks and actions on every message received:
    - Ignores messages sent by bots to prevent the bot from responding to itself or other bots.
    - Checks if the message is from the specified guild (server) by ID. If not, logs the message source and returns.
    - Calls the handle_profanities function to check and act upon messages containing profanities.
    - If the message is in the channel designated for GIFs, it processes the message through handle_gifs_channel.
    - For messages in the recommended channel, checks if the message is considered spam. If not, adds the message to reports.

    Args:
        message (discord.Message): The message object containing data about the received message.
    """
    if message.author.bot:
        return

    if message.guild.id != 1252165373256794185:
        logger.debug(f"Message from {message.guild.name}")
        return

    await handle_profanities(message)

    if message.channel.id == settings.get('gifs_channel_id'):
        await handle_gifs_channel(message)

    # reports
    if message.channel.id == settings.get('recommended_channel_id'):
        if not reports.is_spam(message):
            reports.add_message(message)


@tasks.loop(minutes=1)
async def scheduled_reports_save():
    """
    A scheduled task that saves report data every minute.

    This function logs a message indicating the saving process of report data has started,
    then calls the `save_messages` method of the `reports` object to persistently save the
    collected report messages. This task is scheduled to run every minute to ensure that
    report data is consistently updated and saved.
    """
    logger.info("Saving reports data...")
    reports.save_messages()


# minimum timing : 2 minutes (free plan limitation : 30 messages per hour)
@tasks.loop(minutes=6, hours=24)
async def scheduled_report():
    """
    Generates and sends a daily discussion report to a specified Discord channel.

    This scheduled task runs every 24 hours with a 6-minute offset. It performs the following steps:
    1. Logs into a GPT instance to generate a report prompt.
    2. Sends the generated prompt to the GPT instance and receives a response.
    3. Extracts messages and sentiment analysis from the GPT response.
    4. Compiles a report including the date, sentiment analysis, impactful messages, participants, and warnings.
    5. Sends the compiled report to a specified Discord channel ('moments_channel_id').
    6. Optionally generates a word cloud from the messages and a pie chart for warnings distribution, attaching them to the report.

    Exceptions:
        - Catches and logs any exceptions that occur during the execution.
        - Attempts to close the GPT session gracefully in case of an exception.

    Note:
        - The task relies on external settings for channel IDs and the GPT instance configuration.
        - The 'moments_channel_id' setting determines the target channel for the report.
        - The function uses matplotlib for generating visualizations (word cloud and pie chart).
    """
    global gpt
    try:
        gpt = GPT()
        gpt.login()
        prompt = gpt.generate_report_prompt()
        response = gpt.send_prompt(prompt) if prompt else ""
        if response:
            logger.success(f"ChatGPT response generated.")
            messages = re.findall(r'\[Message \d+\] "(.*?)"', response)
            sentiment_match = re.search(r'Global sentiment: (.*)"', response)
            sentiment = sentiment_match.group(1) if sentiment_match else "No sentiment found"
        else:
            messages = []
            sentiment = "No sentiment found"
        # send the response to moments_channel_id
        moments_channel_id = settings.get('moments_channel_id')
        moments_channel = bot.get_channel(moments_channel_id)
        unique_authors = reports.get_unique_authors()
        unique_authors = [bot.get_user(int(author_id)).name for author_id in unique_authors]
        all_warnings = warnings.get_all_daily_warnings()
        warnings_description = "\n ".join([f"- **{i + 1}**. {bot.get_user(int(user_id)).mention} - {count} warning(s)"
                                            for i, (user_id, count) in enumerate(all_warnings.items())]) \
            if all_warnings else "- No warnings found."

        if moments_channel:
            embed = discord.Embed(title=f":crystal_ball: Daily discussion report",
                                    description=f":date: **Date**: {get_current_date_formatted(separator='/')}\n"
                                                f":bar_chart: **Sentiment**: {sentiment}\n"
                                                f":loudspeaker: **Impactful messages**:\n"
                                                f" - {'\n - '.join(messages) or 'No impactful messages found'}\n"
                                                f":busts_in_silhouette: **Participants**:\n"
                                                f" - {', '.join([f'{author}' for author in unique_authors]) or 'No participants found'}\n"
                                                f":warning: **Warnings**:\n {warnings_description}",
                                    color=0xaa8dd8)
            embed.set_footer(text="MEE7 Daily Report",
                                icon_url=settings.get('icon_url'))
            message = await moments_channel.send(embed=embed)
            # generate word cloud for messages
            if messages:
                wordcloud = WordCloud(width=800, height=400, background_color='white').generate(' '.join(messages))
                plt.figure(figsize=(10, 5))
                plt.imshow(wordcloud, interpolation='bilinear')
                plt.axis('off')
                plt.savefig('wordcloud.png')
                plt.close()
                await message.reply(file=discord.File('wordcloud.png'))
                os.remove('wordcloud.png')

            # generate tree map for warnings
            if all_warnings:
                labels = [bot.get_user(int(user_id)).name for user_id in all_warnings.keys()]
                sizes = [count for count in all_warnings.values()]
                colors = np.random.rand(len(labels), 3)
                plt.figure(figsize=(10, 5))
                plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%')
                plt.axis('equal')
                plt.title('Warnings Distribution')
                plt.savefig('warnings.png')
                plt.close()
                await message.reply(file=discord.File('warnings.png'))
                os.remove('warnings.png')

            # Create and send Excel report
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Daily Report"

            # Set up the headers
            headers = ["Date", "Sentiment", "Messages", "Participants", "Warnings"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)

            # Add the data
            ws.cell(row=2, column=1, value=get_current_date_formatted(separator='/'))
            ws.cell(row=2, column=2, value=sentiment)
            ws.cell(row=2, column=3, value="\n".join(messages))
            ws.cell(row=2, column=4, value=", ".join(unique_authors))
            ws.cell(row=2, column=5, value=warnings_description)

            # Save the Excel file
            excel_file = "daily_report.xlsx"
            wb.save(excel_file)
            await message.reply(file=discord.File(excel_file))
            os.remove(excel_file)
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
    finally:
        try:
            gpt.close()
        except OSError:
            pass


# scheduled_activity_recommendation & scheduled_report must be at least 3 minutes apart to avoid conflicts
@tasks.loop(minutes=3, hours=24)
async def scheduled_activity_recommendation():
    """
    A scheduled task that recommends activities based on the current weather in a specified city.

    This task runs every 24 hours with a 3-minute offset. It fetches the current weather data for a predefined city
    and uses it to generate activity recommendations. These recommendations are then sent to a specific Discord channel
    dedicated to activity suggestions.

    The activity recommendations are generated by sending a prompt to a GPT model, which returns a list of activities
    suitable for the current weather conditions. Each activity includes a name and a URL for more information.

    If the GPT model fails to generate a response or an error occurs during the process, the task will log the error
    and attempt to close the GPT session gracefully. Additionally, if the scheduled report task is not running,
    this task will initiate it to avoid any conflicts.

    Exceptions:
        Logs any exceptions that occur during the execution of the task and attempts to close the GPT session.
    """
    global gpt
    try:
        city = settings.get('city')
        date = datetime.now().strftime('%Y-%m-%d')
        weather_datas = get_weather(city, date)
        gpt = GPT()
        gpt.login()
        prompt = gpt.generate_activity_prompt(weather_datas)
        response = gpt.send_prompt(prompt) if prompt else ""
        if response:
            logger.success(f"ChatGPT response generated.")
            activities = response.split("[")[1].rsplit("]")[0]
            activities = f"[{activities}]"
            activities_json = json.loads(activities)
            activities_names = [activity['activity'] for activity in activities_json]
            urls = [activity['url'] for activity in activities_json]
        else:
            activities_names = []
            urls = []
        # send the response to moments_channel_id
        activity_channel_id = settings.get('activity_channel_id')
        activity_channel = bot.get_channel(activity_channel_id)
        formatted_date = datetime.strptime(date, '%Y-%m-%d').strftime('%m/%d/%Y')
        if activity_channel:
            embed = discord.Embed(title=f":tada: Activity recommendation at {city}",
                                  description=f":date: **Date**: {formatted_date}\n"
                                              f":white_sun_rain_cloud: **Weather**: {weather_datas['weather']}\n"
                                              f":thermometer: **Temperature**: {weather_datas['temperature']}°C",
                                  color=0xfd6ce4)
            embed.set_footer(text="MEE7 Activity Recommendation",
                             icon_url=settings.get('icon_url'))
            for activity_name, url in zip(activities_names, urls):
                embed.add_field(name=f":placard: {activity_name}", value=f":link: [Read more]({url})", inline=False)
            await activity_channel.send(embed=embed)
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
    finally:
        try:
            gpt.close()
            if not scheduled_report.is_running():
                # Start the scheduled report task after the activity recommendation task to avoid conflicts
                start_scheduled_report.start()
        except OSError:
            pass


@tasks.loop(minutes=3)
async def start_scheduled_report():
    """
    A scheduled task that attempts to start the scheduled report task.

    This task runs every 3 minutes. On its first loop, it simply logs a message indicating a delay
    before starting the scheduled report task. This is to ensure there's a delay before the task starts,
    allowing for any necessary setup or conditions to be met.

    On subsequent loops, if the scheduled report task is not already running, it starts the task and then
    stops itself to prevent further execution.
    """
    # Do not start the task on the first loop, so that we add a delay
    if start_scheduled_report.current_loop == 0:
        logger.debug("Delaying the start of the scheduled report task, starting in 3 minutes...")
        return
    if not scheduled_report.is_running():
        logger.info("Starting the scheduled report task...")
        scheduled_report.start()
        logger.success("Scheduled report task started.")
        start_scheduled_report.stop()

@tasks.loop(hours=1)
async def scheduled_recommendation():
    """
    This function is a task that runs every hour. Its purpose is to send a recommendation message to a specific channel.

    The function first retrieves the channel using the channel_id. If the channel exists, it sends a message
    indicating that it's analyzing and recommending content. It then calls the `analyze_and_recommend` function with
    the bot and channel_id as arguments to get the recommendation. The recommendation is then sent to the channel.

    This function doesn't take any arguments and doesn't return anything.
    """
    recommended_channel_id = settings.get('recommended_channel_id')
    recommended_channel = bot.get_channel(recommended_channel_id)
    channel_id = settings.get('recommendations_channel_id')
    channel = bot.get_channel(channel_id)

    if channel and recommended_channel:
        embed = discord.Embed(title=":alarm_clock: Scheduled Recommendation",
                                description=f":mag_right: Analyzing and recommending content in {recommended_channel.mention}...",
                                color=0xe6e7e8)
        embed.set_footer(text="MEE7 Recommendation System",
                            icon_url=settings.get('icon_url'))
        message = await channel.send(embed=embed)
        recommendation = await generate_recommendations(bot, recommended_channel, recommended_channel_id, discord, settings.get('icon_url'))

        await message.reply(embed=recommendation)


@bot.command(name="recommend", description="Recommends content based on recent discussions")
async def recommend(ctx, channel: discord.TextChannel):
    """
    This function is a command handler for the 'recommend' command.

    It takes two arguments: - ctx: The context in which the command was called. This includes information about the
    message, the channel, and the user who called the command. - channel_id: The ID of the channel for which to
    recommend content. This should be a string.

    The function first checks if the channel_id is a digit. If it is, it converts it to an integer and checks if a
    channel with that ID exists. If the channel exists, it calls the `analyze_and_recommend` function with the bot
    and channel_id as arguments to get the recommendation. The recommendation is then sent as a response to the
    command. If the channel doesn't exist, it sends a response indicating that the channel was not found for
    analysis. If the channel_id is not a digit, it sends a response indicating that the channel ID is invalid.

    This function doesn't return anything.
    """
    channel_id = channel.id
    channel = bot.get_channel(channel_id)
    if channel:
        recommendation = await generate_recommendations(bot, channel, channel_id, discord, settings.get('icon_url'))
        await ctx.respond(embed=recommendation)
    else:
        await ctx.respond("Channel not found for analysis.")


@bot.command(name="warnings", description="Displays the warnings for a user or all users")
async def display_warnings(ctx, user: discord.User = None):
    """
    This function is a command handler for the 'warnings' command.

    Args: ctx (discord.Context): The context in which the command was called. user (discord.User, optional): The user
    for whom to display warnings. If not provided, warnings for all users are displayed.

    The function first checks if a user was provided. If a user was provided, it retrieves the warnings for that user
    and sends a response with the number of warnings. If no user provided, it retrieves the warnings for all
    users. If there are any warnings, it creates a description string with the warnings for each user and sends a
    response with an embed message containing the warnings. If there are no warnings, it sends a response indicating
    that no warnings were found.

    This function doesn't return anything.
    """
    if user:
        await ctx.respond(f"{user.mention} has {warnings.get_user_warnings(user.id)} warning(s).")
    else:
        all_warnings = warnings.get_all_warnings()
        if all_warnings:
            description = "\n".join([f"**{i + 1}**. {ctx.guild.get_member(int(user_id)).mention} - {count} warning(s)"
                                     for i, (user_id, count) in enumerate(all_warnings.items())])
        else:
            description = "No warnings found."
        embed = discord.Embed(title=f":warning: Warnings Summary of {ctx.guild.name}",
                              color=discord.Color.red(),
                              description=description)
        embed.set_footer(text="MEE7 Warning System",
                         icon_url=settings.get('icon_url'))
        await ctx.respond(embed=embed)


def load_user_icals(directory='user_icals'):
    """
    Loads user iCal data from JSON files within a specified directory.

    This function scans a given directory for JSON files, each expected to contain iCal data for a user. It constructs
    a dictionary mapping user IDs (extracted from the filenames) to the full path of their respective JSON file. If the
    specified directory does not exist, it is created.

    Args:
        directory (str): The directory to scan for JSON files. Defaults to 'user_icals'.

    Returns:
        dict: A dictionary where keys are user IDs (str) and values are the paths (str) to their iCal JSON files.
    """
    users_icals = {}
    if not os.path.exists(directory):
        os.makedirs(directory)
    for filename in os.listdir(directory):
        if filename.endswith('.json'):
            user_id = filename.split('.')[0]
            users_icals[user_id] = os.path.join(directory, filename)
    return users_icals


user_icals = load_user_icals()


@bot.command(name="register_ical", description="Register your iCal file for availability checks")
async def register_ical(ctx, url: discord.Option(discord.SlashCommandOptionType.string)):
    """
    Registers a user's iCal file for availability checks by downloading it and associating it with their Discord ID.

    This command is triggered by a Discord slash command. It first ensures that a temporary directory exists for storing
    iCal files. Then, it downloads the iCal file from the provided URL to a temporary file. After downloading,
    it registers
    the user's iCal file by associating the temporary file path with the user's Discord ID and name in a global
    dictionary.
    Finally, it performs an initial planning operation (e.g., checking availability) and responds to the user indicating
    successful registration.

    Args:
        ctx: The context of the command, which includes information about the channel, guild, and user who invoked.
        url: A string representing the URL of the iCal file to be downloaded and registered.

    Returns:
        None. The function's primary side effects are downloading a file, updating a global dictionary, performing an
        initial planning operation, and sending a response message to the Discord channel.
    """
    ensure_temp_dir()  # Ensure the temporary directory for iCal files exists.
    temp_file_path = os.path.join(TEMP_DIR, 'temp.ics')  # Define the path for the temporary iCal file.
    await download_ical(url, temp_file_path, ctx)  # Download the iCal file from the provided URL.
    await register_user_ical(ctx.author.id, ctx.author.name, temp_file_path,
                             user_icals)  # Register the user's iCal file.
    await planning(ctx)  # Perform an initial planning operation with the newly registered iCal file.
    await ctx.respond(f":white_check_mark: Your iCal file has been registered successfully.")  # Respond to the user.


@bot.command(name="availability", description="Displays the availabilities of all persons in the Discord server")
async def availability(ctx):
    """
    A Discord bot command to display the availability of all members in the server.

    This command retrieves all non-bot members from the guild (server) and presents a selection menu to the command invoker.
    Upon selection of a user, it fetches the user's availability from a JSON file named after the user's ID and displays
    it in the Discord channel.

    Args:
        ctx: The context under which the command is executed. Contains information and methods related to the command
        invocation.

    Returns:
        None. This function operates by sending messages to a Discord channel.
    """
    # Retrieve all non-bot members from the guild
    users = ctx.guild.members
    users = [user for user in users if not user.bot]

    # If no users are found, send a response and return
    if not users:
        await ctx.respond("No users found in the Discord server.")
        return

    # Create selection options for each user
    options = [discord.SelectOption(label=member.display_name, value=str(member.id)) for member in users]

    # Create a selection menu with the user options
    select = Select(placeholder="Choose a user", options=options)

    async def select_callback(interaction):
        """
        Callback function for when a user is selected from the dropdown.

        Fetches the selected user's availability from a JSON file and sends it as an embed in the Discord channel.

        Args:
            interaction: The interaction object representing the user's selection.

        Returns:
            None. Sends messages to a Discord channel.
        """
        # Retrieve the selected user's ID from the selection
        user_id = int(select.values[0])
        file_path = f"user_icals/{user_id}.json"

        # Check if the file exists
        if not os.path.exists(file_path):
            await interaction.response.send_message(
                f"No availability data found for {interaction.guild.get_member(user_id).display_name}.")
            return

        # Show typing indicator while processing
        async with ctx.typing():
            try:
                # Fetch and display the selected user's availability
                embeds = await is_everyone_available(ctx, file_path)
                for embed in embeds:
                    await interaction.response.send_message(embed=embed)
            except Exception as e:
                await interaction.response.send_message(f"An error occurred while retrieving the data: {str(e)}")

    # Assign the callback to the selection
    select.callback = select_callback
    # Create a view and add the selection menu to it
    view = View()
    view.add_item(select)

    # Send the selection menu to the Discord channel
    await ctx.respond("Select a user to view their availability:", view=view)


def aggregate_weekly_events(directory='user_icals'):
    """
    Aggregates weekly events for all users from JSON files within a specified directory.

    This function scans a directory for JSON files, each representing a user's event data. For each file,
    it extracts the user's ID and iCal content, then parses the iCal content to determine the user's availability
    for the current week. The availability data is aggregated into a dictionary, keyed by user ID, with each value
    being another dictionary mapping ISO-formatted dates to availability information.

    Args:
        directory (str): The directory to scan for user JSON files. Defaults to 'user_icals'.

    Returns:
        dict: A dictionary where each key is a user ID (str) and each value is a dictionary. The value dictionary
              maps ISO-formatted dates (str) to availability information (dict), which indicates the user's
              availability for morning, afternoon, and evening of each day in the current week.
    """
    aggregated_events = {}
    current_week_start = get_current_week_start()

    for filename in os.listdir(directory):
        if filename.endswith('.json'):
            with open(os.path.join(directory, filename), 'r') as json_file:
                user_data = json.load(json_file)
                user_id = str(user_data["user_id"])
                ical_content = user_data.get("ical_content", "")

                if ical_content:
                    events = parse_ical_content(ical_content)
                    if events is None:
                        continue
                    week_events = check_availability(events, current_week_start)

                    week_events_str_keys = {day.isoformat(): availability for day, availability in week_events.items()}

                    aggregated_events[user_id] = week_events_str_keys

    return aggregated_events


async def planning(ctx):
    """
    Asynchronously plans and aggregates weekly events for all non-bot users in a Discord server.

    Args:
        ctx: The context from which this function is called. Provides access to Discord server information and
             methods for responding to events.

    Returns:
        None. This function primarily interacts with the Discord API to send messages and does not return any value.
    """
    ensure_temp_dir()  # Ensure the temporary directory for storing files exists.
    users = ctx.guild.members  # Retrieve all members from the guild (Discord server).
    users = [user for user in users if not user.bot]  # Filter out bot users.

    if not users:
        await ctx.respond("No users found in the Discord server.")  # Respond if no non-bot users are found.
        return

    aggregated_events = aggregate_weekly_events()  # Aggregate weekly events for the users.

    if not aggregated_events:
        await ctx.respond("No events found for the current week.")  # Respond if no events are aggregated.
        return

    combined_json = json.dumps(aggregated_events,
                               indent=4)  # Convert the aggregated events into a formatted JSON string.

    # Write the JSON string to a file for storage or further processing.
    if not os.path.exists("events_state"):
        os.makedirs("events_state")
    with open("events_state/aggregated_events.json", "w") as json_file:
        json_file.write(combined_json)


@bot.command(name="display_common_availability",
             description="Displays common availability of all users for the current week")
async def display_common_availability(ctx):
    """
    Displays the common availability of all users in the Discord server for the current week.

    This command aggregates the availability of all users for the current week and presents it in an embed message.
    The availability is categorized into three time slots: morning, afternoon, and evening. For each day of the
    current week, the command lists the users available during these time slots. Days when users are available
    across all time slots are highlighted with a star symbol.

    Args:
        ctx: The context under which the command is executed. Contains information about the guild, channel, and user.

    Returns:
        An embed message sent to the channel from which the command was invoked. The message contains the common
        availability of users for each day of the current week, categorized into morning, afternoon, and evening.
    """
    weekly_events = aggregate_weekly_events()
    all_users = ctx.guild.members

    time_slots = ["morning", "afternoon", "evening"]

    common_availability = {}

    for user in all_users:
        user_id = str(user.id)
        if user_id in weekly_events:
            user_name = user.display_name
            for day, times in weekly_events[user_id].items():
                if day not in common_availability:
                    common_availability[day] = {slot: [] for slot in time_slots}
                for time_slot in time_slots:
                    if times.get(time_slot) is True:
                        common_availability[day][time_slot].append(user_name)

    max_availability = {slot: 0 for slot in time_slots}
    for slots in common_availability.values():
        for time_slot, users in slots.items():
            max_availability[time_slot] = max(max_availability[time_slot], len(users))

    embed = discord.Embed(title="Common Availability for All Users", color=discord.Color.green())

    dates_with_stars = []

    for day, slots in common_availability.items():

        date_obj = datetime.strptime(day, '%Y-%m-%d')
        french_day = date_obj.strftime('%A %d %B')
        french_day = french_day.capitalize()

        morning_users = ", ".join(slots["morning"])
        afternoon_users = ", ".join(slots["afternoon"])
        evening_users = ", ".join(slots["evening"])

        day_star = ""
        if all(len(slots[slot]) > 0 for slot in time_slots):
            day_star = "⭐"
            dates_with_stars.append(day)

        availability_str = (
            f"```\nMorning : {morning_users}\n"
            f"Afternoon : {afternoon_users}\n"
            f"Evening : {evening_users}\n```"
        )
        embed.add_field(name=f"{french_day} {day_star}", value=availability_str, inline=False)
    await ctx.respond(embed=embed)


def display_best_days():
    """
    Identifies and returns the days when all users have common availability in all time slots.

    This function aggregates weekly events for all users, iterates through each user's events,
    and builds a dictionary mapping each day to the users available during morning, afternoon, and evening slots.
    It then identifies days when there is at least one user available in each time slot and returns these days.

    The function assumes the existence of a `aggregate_weekly_events` function that aggregates events for all users
    and a `get_display_name_from_id` function that returns a user's display name given their user ID.

    Returns:
        list: A list of dates (as strings) where all users have common availability in all time slots.
    """
    weekly_events = aggregate_weekly_events()
    all_users = []

    # Loop through the files in the 'user_icals' directory to get user IDs
    for filename in os.listdir('user_icals'):
        if filename.endswith('.json'):
            with open(os.path.join('user_icals', filename), 'r') as json_file:
                user_data = json.load(json_file)
                user_id = str(user_data["user_id"])
                all_users.append(user_id)

    time_slots = ["morning", "afternoon", "evening"]

    common_availability = {}

    # Iterate over all users to build the common availability dictionary
    for user_id in all_users:
        if user_id in weekly_events:
            # Assuming you have a way to get the display name from the user_id
            user_name = get_display_name_from_id(user_id)  # Placeholder function
            for day, times in weekly_events[user_id].items():
                if day not in common_availability:
                    common_availability[day] = {slot: [] for slot in time_slots}
                for time_slot in time_slots:
                    if times.get(time_slot):
                        common_availability[day][time_slot].append(user_name)

    max_availability = {slot: 0 for slot in time_slots}
    for slots in common_availability.values():
        for time_slot, users in slots.items():
            max_availability[time_slot] = max(max_availability[time_slot], len(users))
    dates_with_stars = []
    for day, slots in common_availability.items():
        if all(len(slots[slot]) > 0 for slot in time_slots):
            dates_with_stars.append(day)
    return dates_with_stars


def get_display_name_from_id(user_id):
    """
    Generates a display name for a user based on their user ID.

    This function simply formats the user ID into a string that starts with "User_"
    followed by the user ID. It's a straightforward way to generate a unique display name
    for users based on their ID.

    Args:
        user_id: The unique identifier for the user.

    Returns:
        A string that represents the display name of the user.
    """
    return f"User_{user_id}"


def get_current_week_start():
    """
    Calculate the start date of the current week based on the Europe/Paris timezone.

    This function determines the current date and time in the 'Europe/Paris' timezone,
    then calculates the start of the week (Monday) by subtracting the current weekday
    number from the current date. The weekday function returns 0 for Monday through 6 for Sunday,
    aligning with the ISO standard for the first day of the week.

    Returns:
        datetime.date: The date representing the start (Monday) of the current week.
    """
    return datetime.now(pytz.timezone('Europe/Paris')).date() - timedelta(
        days=datetime.now(pytz.timezone('Europe/Paris')).weekday())


@bot.command(name="add_streamer", description="Adds a streamer to the list of streamers to check")
@commands.has_permissions(administrator=True)
async def add_streamer(ctx, streamer: discord.Option(discord.SlashCommandOptionType.string)):
    """
    This function is a command handler for the 'add_streamer' command.

    It takes two arguments:
    - ctx: The context in which the command was called.
    - streamer: The name of the streamer to add to the list of streamers to check.

    The function first validates the streamer name using the `validate_streamer` function. If the streamer is valid,
    it adds the streamer to the list of streamers to check and sends a success message. If the streamer is invalid,
    it sends an error message.
    """
    streamer = streamer.lower().replace(" ", "")
    if await validate_streamer(streamer, append=True):
        settings.add_streamer(streamer)
        await ctx.respond(f":white_check_mark: {streamer} has been added to the list of streamers to check.")
    else:
        await ctx.respond(f":x: {streamer} is not a valid Twitch username.")


@bot.command(name="top10messages", description="Displays the top 10 users who sent the most messages today.")
async def top10messages(ctx, bots: discord.Option(discord.SlashCommandOptionType.boolean) = False):
    """
    Asynchronously handles the 'top10messages' command within Discord.

    This command calculates and displays the top 10 users who have sent the most messages in the server for the current day.
    It can optionally include messages sent by bots.

    Args:
        ctx (discord.Context): The context in which the command was called. This includes information about the message,
                               the channel, and the user who called the command.
        bots (discord.Option): A boolean option provided by the user indicating whether to include bot messages in the count.
                               Defaults to False, excluding bot messages from the count.

    The function performs the following steps:
    1. Sends an initial response indicating that the calculation is in progress.
    2. Initializes a dictionary to keep track of message counts per user.
    3. Iterates through all text channels in the server, counting messages sent by each user after midnight of the current day.
       If 'bots' is False, messages sent by bots are excluded.
    4. Sorts the users by their message count in descending order and selects the top 10.
    5. If there are no messages found for the current day, sends an embed message indicating so.
    6. Otherwise, generates a bar graph displaying the usernames and their corresponding message counts.
    7. Saves the graph to a file named 'top10messages.png', sends it in an embed message, and then deletes the file.
    """
    await ctx.respond("Calculating, this may take a moment...")

    message_counts = {}
    today = datetime.now(timezone.utc).date()

    # Iterate through all channels in the server
    for channel in ctx.guild.text_channels:
        async for message in channel.history(limit=None,
                                             after=datetime.combine(today, datetime.min.time(), tzinfo=timezone.utc)):
            if message.author.bot and not bots:
                continue
            if message.author.id in message_counts:
                message_counts[message.author.id] += 1
            else:
                message_counts[message.author.id] = 1

    # Sort users by message count
    sorted_counts = sorted(message_counts.items(), key=lambda item: item[1], reverse=True)
    top10 = sorted_counts[:10]

    if not top10:
        # No messages found today
        embed = discord.Embed(title=f"No messages found today ({today})",
                              description="No data available for top 10 users by message count.",
                              color=discord.Color.blue())
        embed.set_footer(text="MEE7 Stats", icon_url=settings.get('icon_url'))
        await ctx.respond(embed=embed)
        return

    # Extract usernames and message counts
    user_names = []
    message_numbers = []
    for user_id, count in top10:
        user = await bot.fetch_user(user_id)
        user_names.append(f"{user.name}#{user.discriminator}" if user.discriminator != "0" else user.name)
        message_numbers.append(count)

    # Generate the graph
    plt.figure(figsize=(10, 5))
    plt.bar(user_names, message_numbers, color='skyblue')
    for i, v in enumerate(message_numbers):
        plt.text(i, v + 0.5, str(v), ha='center', va='bottom')
    plt.xlabel('Users')
    plt.ylabel('Message Count')
    plt.title(f"Top 10 Users by Message Count (Today, {today})")
    plt.xticks(rotation=45)
    plt.tight_layout()

    # Save the graph to a file
    plt.savefig('top10messages.png')
    plt.close()

    # Send the graph on Discord in an embed
    embed = discord.Embed(title=f"Top 10 Users by Message Count (Today, {today})",
                          color=discord.Color.blue())
    embed.set_image(url="attachment://top10messages.png")
    embed.set_footer(text="MEE7 Stats", icon_url=settings.get('icon_url'))
    await ctx.respond(embed=embed, file=discord.File('top10messages.png'))
    os.remove('top10messages.png')


@bot.command(name='sb-ultras', description='Display the list of ultra abilities')
async def sb_ultras(ctx, character: Option(str, "The character name (Archer Queen, Barbarian...)", required=False)):
    """
    Handles the 'sb-ultras' command in Discord. It displays a list of ultra abilities for characters in the game.

    This command can be used with or without specifying a character name. If a character name is provided, the command
    displays the ultra abilities for that specific character. If no character name is provided, it displays the ultra
    abilities for all characters.

    Args:
        ctx: The context under which the command is executed. Contains information about the guild, channel, and user.
        character: An optional argument. The name of the character for which to display ultra abilities. If not provided,
                   the command displays ultra abilities for all characters.

    The function reads the ultra abilities data from a JSON file (`abilities.json`), which contains the abilities for
    each character. The character names are keys in this JSON file. If a character name is provided, it converts the
    name to lowercase and replaces spaces with hyphens to match the keys in the JSON file. If the character is found,
    it displays the abilities for that character. Otherwise, it sends a message indicating the character was not found.
    """
    with open("src/ft/bonus/squadbusters/abilities.json", "r") as file:
        abilities_data = json.load(file)
    characters = [character for character in abilities_data.keys() if character != 'description']

    if character:
        character = character.lower().replace(" ", "-")
        if character in abilities_data:
            start_index = characters.index(character)
        else:
            await ctx.respond(f"Character **{character}** not found.")
            return
    else:
        start_index = 0

    view = NavigationView(characters, abilities_data, start_index)
    embed = view.update_embed()
    await ctx.respond(embed=embed, view=view)



@bot.slash_command(name="raids", description="Display the list of raids and their steps")
async def raids(ctx):
    """
    Handles the '/raids' slash command in Discord. It allows users to select a raid from a predefined list and then
    choose a specific step of the raid to view detailed information about it.

    Args:
        ctx: The context under which the slash command is executed. Contains information about the guild, channel,
             and user.

    This function first presents a selection menu of available raids to the user. Once a raid is selected, another
    selection menu is presented with the steps of the chosen raid. The user can then select a specific step to view
    detailed information about it, which is displayed in an embed message.
    """
    raids = ["Vow of the Disciple"]
    raid_options = [discord.SelectOption(label=raid, value=raid) for raid in raids]

    raid_select = Select(placeholder="Choose a raid...", options=raid_options)

    async def raid_select_callback(interaction):
        """
        Callback function for when a raid is selected from the raid selection menu.

        Args:
            interaction: The interaction object representing the user's selection.

        This function retrieves the selected raid, validates it, and then presents another selection menu to the user
        with the steps available for the selected raid. If the selected raid is not found, an error message is sent.
        """
        selected_raid = raid_select.values[0]

        if selected_raid not in raids:
            await interaction.response.send_message("Invalid option.", ephemeral=True)
            return

        raid_path = os.path.join(os.getcwd(), "src/ft/bonus/destiny/", selected_raid)
        if not os.path.exists(raid_path):
            await interaction.response.send_message(f"Folder {raid_path} unreachable.", ephemeral=True)
            return

        etapes = [etape.split('.')[0] for etape in os.listdir(raid_path) if etape.endswith('.txt')]
        etapes = [etape.replace("_", " ") for etape in etapes]
        etape_options = [discord.SelectOption(label=etape, value=etape) for etape in etapes]

        etape_select = Select(placeholder="Choose a step...", options=etape_options)

        async def etape_select_callback(interaction):
            """
            Callback function for when a raid step is selected from the step selection menu.

            Args:
                interaction: The interaction object representing the user's selection.

            This function retrieves the selected step, validates it, and then reads the corresponding information
            from a text file. The information is then displayed to the user in an embed message. If the selected step
            is not found, an error message is sent.
            """
            selected_etape = etape_select.values[0]

            file_path = os.path.join(raid_path, f'{selected_etape.lower().replace(" ", "_")}.txt')

            if not os.path.isfile(file_path):
                await interaction.response.send_message(f"File {file_path} unreachable.", ephemeral=True)
                return

            try:
                with open(file_path, 'r', encoding='utf-8') as file:
                    content = file.read()
            except Exception as e:
                await interaction.response.send_message(f"Error when reading file : {e}", ephemeral=True)
                return

            embed = discord.Embed(title=f"{selected_raid} - {selected_etape}", description=content, color=discord.Color.blue())
            embed.set_image(url=content.split('\n')[0])

            await interaction.response.send_message(embed=embed, ephemeral=False)

        etape_select.callback = etape_select_callback

        etape_view = View()
        etape_view.add_item(etape_select)

        await interaction.response.send_message("Select a step :", view=etape_view, ephemeral=False)

    raid_select.callback = raid_select_callback

    raid_view = View()
    raid_view.add_item(raid_select)

    await ctx.respond("Select a raid  :", view=raid_view, ephemeral=False)


setup_commands(bot)
bot.run(DISCORD_BOT_TOKEN)
